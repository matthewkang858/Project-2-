#!/usr/bin/env python3
"""Build the "QA - Claude" reviewer tab in the COUR4 QA workbook.

Trace-grade: everything below is derived from one complete console-bot trace
of the live Qualtrics survey (run-001, 109 screens, one path to completion)
and cross-checked, question by question, against the parsed v15 questionnaire
spec (surveys/cour4/outline-spec.json).

Boolean columns follow the JB/MK convention: TRUE = affirmatively verified /
applies-and-passes on this pass; blank/FALSE = not positively confirmed (may
be fine, may be flagged — see the comment). Columns filled here:
  D (Wording matches)  - live stem + answer text read against v15
  E (Formatting)       - live bold/underline compared to v15 emphasis
  F (SP vs MP)         - live control type vs v15 SP/MP
C, G, H, I, J are left blank: a single path cannot certify skippability,
display-rule correctness, randomization/anchor order, or exclusive logic.

The template's B column uses the LIVE Qualtrics numbering (it runs to Q119),
so trace question numbers map straight onto rows; the extra live Q82 shifts
live Q83+ to questionnaire Q82+ (see the Q82 comment).
"""
import re
import openpyxl
from openpyxl.styles import PatternFill

# red fill for boolean grid cells that encode a detected problem
RED = PatternFill(fill_type="solid", fgColor="FFFF4C4C")

SRC = "/root/.claude/uploads/c008ad0f-2509-568b-91ec-bf12d8f404a9/66fad3d8-COUR4_Consumer_Survey_QA.xlsx"
OUT = "/tmp/claude-0/-home-user-Project-2-/c008ad0f-2509-568b-91ec-bf12d8f404a9/scratchpad/COUR4_Consumer_Survey_QA_Claude.xlsx"

# --- verified from trace vs spec (see build notes) ---
# live Q-numbers reached on this path
REACHED = [1,2,3,4,5,6,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,25,26,27,
    28,29,31,33,34,35,36,37,38,40,41,42,43,55,56,57,60,61,62,63,64,65,69,70,71,
    72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,96,99,
    101,102,109,110,111,112,113,114,115,116,117,118,119]
# SP/MP verified correct on every non-text reached question (0 mismatches / 80 checked)
F_OK = [n for n in REACHED if n not in (14,34,35,55)]
# live bold+underline set matched v15 emphasis exactly
E_OK = [1,2,3,4,5,8,9,10,11,14,16,17,18,19,20,21,26,27,29,31,34,37,40,56,69,70,
    71,72,73,74,75,76,79,80,83,85,87,88,90,93,96,99,101,102,110,111,112,113,114,
    115,117,118,119]
# wording / answer-text discrepancy vs v15 (these get D = FALSE)
# NOTE: the checkbox "blank option" flags (Q9/Q10/Q25/Q27/Q92/Q94) were REMOVED.
# On those, the trace captured the option label as the literal word "Selected" (the
# checkbox state) with empty emphasis, which is the bot's fall-back when it cannot
# read an option's label element -- a capture artifact, not a real blank render.
# Corroboration: JB/MK (two human reviewers) did not flag any of them, and confirmed
# only the *radio* bare-number blanks (Q6="2", Q87="4"). Kept: real text/typo findings
# and the human-corroborated radio blanks.
D_BAD = [4,6,28,36,38,40,41,42,57,61,64,65,82,87,109,112]

COMMENTS = {
    4: (  # Q4
        'Minor wording mismatch in the last option.\n'
        'The live survey shows "Prefer not to answer"; the questionnaire (and Q3, Q110) '
        'uses "Prefer not to say". Pick one and make it consistent.'
    ),
    6: (  # Q6
        'Two problems in the answer list (both also caught by JB and MK).\n'
        '1. Option b displays with no text at all — it comes through as just "2".\n'
        '2. Option c reads "I had did not have any control...". Drop the extra "had": '
        '"I did not have any control over which online learning platform I used".'
    ),
    28: (  # Q28
        'Stem wording differs from the questionnaire.\n'
        'The live stem reads "...which of the following topics are you interested in studying '
        'through an online learning platform?", but the answer grid is a Beginner / '
        'Intermediate / Advanced level matrix. The questionnaire stem asks "...what level are '
        'you most interested in studying for each of the following areas...". Confirm which '
        'stem is intended — as shown, the stem and the answer grid don\'t line up.'
    ),
    36: (  # Q36
        'Typo in option b (also caught by JB and MK).\n'
        'Live: "...without setting a specific or plan budget". Questionnaire: "...without '
        'setting a specific budget or plan". The last two words are swapped.'
    ),
    38: (  # Q38
        'Stem is missing a word — and this question gates most of the pricing/concept section '
        'that follows (Q40+).\n'
        'Live reads "...which of the following would you be interested in?"; the questionnaire '
        'reads "...interested in purchasing?". The word "purchasing" is missing. (JB '
        'separately flagged the anchor settings on this question.)'
    ),
    40: (  # Q40
        'Typo in option c.\n'
        'Live reads "Somewhat likely"; the rest of the scale is "...likely to pay" '
        '(Extremely / Very / Slightly / Not at all likely to pay). Should be "Somewhat likely to pay".'
    ),
    41: (  # Q41
        'Please confirm on screen: piped price may be missing its "$".\n'
        'In the trace the stem came through as "Would you be willing to pay 40.00 for this '
        'individual course..." — no "$" in front of the number, the same way at every price '
        'point, and likewise on Q61 and Q65. The "$" renders fine elsewhere (e.g. Q4), so '
        'this looks real, but worth a quick visual check since it is a piped value.'
    ),
    42: (  # Q42
        'Please confirm on screen: option b may display blank.\n'
        'The automated trace read option b as just "2" — the same signature as the blank '
        'options JB/MK confirmed on Q6 and Q87. The intended text is "I would not purchase a '
        'course with limited access...". Q42 is display-gated, so it may not have been on the '
        'path JB/MK reviewed; worth a look.'
    ),
    57: (  # Q57
        'Typo in the last option.\n'
        'Live reads "Not at interested". Should be "Not at all interested" to match the scale '
        'used elsewhere (e.g. Q11).'
    ),
    61: (  # Q61
        'Please confirm on screen: piped price may be missing its "$" (same as Q41).\n'
        'Trace stem: "...willing to pay 87.00 for a verified skill assessment..." with no "$".'
    ),
    64: (  # Q64
        'Typo in option b.\n'
        'Live reads "Very likely likely to pay" — "likely" is duplicated. Should be '
        '"Very likely to pay".'
    ),
    65: (  # Q65
        'Please confirm on screen: piped price may be missing its "$" (same as Q41).\n'
        'Trace stem: "...willing to pay 50.00 per month for this subscription..." with no "$".'
    ),
    82: (  # Q82  (extra live question, no questionnaire equivalent)
        'Confirms JB/MK: this question is not in the questionnaire.\n'
        'The live survey has an extra single-select here ("When you think about what makes a '
        'learning experience worth paying for, which of the following matters most to you?"), '
        'with 3 options on non-sequential value codes (1 / 4 / 5 — itself a sign options were '
        'deleted). It does not appear in v15, and it pushes every downstream question one '
        'number higher in the live survey than in the questionnaire (live Q119 = '
        'questionnaire Q118).'
    ),
    87: (  # Q87
        'Confirms JB/MK: option b displays blank.\n'
        'It comes through as just "4" on screen; the missing option is "I would want partial '
        'access (i.e., some features or content unlocked while others are gated)...". Note the '
        'option value codes also jump (1 / 4 / 5 / 6), so only four options are present — '
        'confirm none were dropped.'
    ),
    109: (  # Q109  (live Q109 = questionnaire Q108)
        'Confirms JB/MK: answer list and stem don\'t match the questionnaire.\n'
        'Stem: live says "highest degree of education", the questionnaire says "highest level '
        'of education". Options: the live list is 10 items in census wording ("Some high '
        'school, no diploma"; "High school graduate, diploma or equivalent (e.g. GED)"; "Some '
        'college credit, no degree"; ...), while the questionnaire lists 9 in different '
        'wording ("High school diploma or GED"; "Some college, no degree").'
    ),
    112: (  # Q112  (live Q112 = questionnaire Q111)
        'Confirms JB/MK: one answer option is missing.\n'
        'The live survey has 9 options; the questionnaire has 10. The missing one is '
        '"Prefer not to say".'
    ),
}

SCOPE_NOTE = (
    "QA - Claude (independent third-reviewer pass).\n"
    "Built from one complete automated trace of the LIVE Qualtrics survey (one path, run to "
    "completion, 109 screens), then compared question-by-question against the v15 "
    "questionnaire.\n\n"
    "Boolean columns use the same convention as the JB/MK tabs: TRUE = checked and verified "
    "on this pass; FALSE = not positively confirmed (not necessarily wrong — see the "
    "comment). Filled here:\n"
    "  • Wording (D): live stem + answer text read against v15. TRUE where they match; FALSE "
    "where a discrepancy was found (see comment).\n"
    "  • Formatting (E): live bold/underline compared to v15 emphasis. TRUE only where they "
    "matched exactly; FALSE where not confirmed — a manual formatting pass is still worth "
    "doing.\n"
    "  • SP vs MP (F): live control type vs v15. TRUE on all reached questions — no SP/MP "
    "mismatch was found.\n"
    "Cells shaded RED are the flagged ones — a boolean that encodes a detected problem (all "
    "in the Wording column on this pass); each has a matching comment.\n\n"
    "Left FALSE on purpose (a single path cannot judge these — see the JB/MK multi-run tabs): "
    "'Question not skippable' (C), 'Display logic' (G), 'Randomization/anchor' (H), "
    "'Exclusive logic' (I), 'Other logic' (J).\n"
    "31 display-gated questions were not reached on this path and are left FALSE throughout.\n\n"
    "Reliability note: this tab reports only blank/no-text answer options that show as a bare "
    "value number on a single-select AND were independently confirmed by a human reviewer "
    "(Q6, Q87). Automated 'blank option' flags on multi-selects were dropped as capture "
    "artifacts — the tool logged the option's checkbox state ('Selected') rather than reading "
    "its label, which is not evidence the option is blank on screen. Two items (Q41/Q61/Q65 "
    "missing '$', Q42 option b) are marked 'confirm on screen' rather than asserted."
)


def main():
    wb = openpyxl.load_workbook(SRC)
    tmpl = wb["QA Template"]
    ws = wb.copy_worksheet(tmpl)
    ws.title = "QA - Claude"
    wb.move_sheet(ws, offset=-(wb.sheetnames.index("QA - Claude") - (wb.sheetnames.index("QA - MK") + 1)))

    # map live Q-number -> row (template B column holds live numbering)
    row_of = {}
    for r in range(4, ws.max_row + 1):
        b = ws.cell(r, 2).value
        if isinstance(b, int):
            row_of[b] = r
        elif isinstance(b, str) and re.fullmatch(r"\d+", b.strip()):
            row_of[int(b.strip())] = r

    ws["K1"] = SCOPE_NOTE

    C = {"D": 4, "E": 5, "F": 6}  # column indices
    for n in REACHED:
        r = row_of.get(n)
        if not r:
            continue
        ws.cell(r, C["D"]).value = (n not in D_BAD)   # wording
        ws.cell(r, C["E"]).value = (n in E_OK)        # formatting
        ws.cell(r, C["F"]).value = (n in F_OK)        # SP vs MP
        # red-highlight the boolean cell(s) that encode an actual finding.
        # Only the Wording (D) column carries detected problems on this pass;
        # E/F FALSE mean "not confirmed", not a defect, so they stay unfilled.
        if n in D_BAD:
            ws.cell(r, C["D"]).fill = RED

    for n, text in COMMENTS.items():
        r = row_of.get(n)
        if r:
            ws.cell(r, 11).value = text  # column K

    wb.save(OUT)
    print("wrote", OUT)
    print("comments:", len(COMMENTS), "| D-flags:", len(D_BAD),
          "| E=TRUE:", len(E_OK), "| F=TRUE:", len(F_OK))


if __name__ == "__main__":
    main()

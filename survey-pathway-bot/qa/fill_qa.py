#!/usr/bin/env python3
"""Fill the QA workbook's per-question columns from the questionnaire spec and,
when available, a captured run of the live survey.

    python3 qa/fill_qa.py --workbook OUT27_B2B_Survey_QA_Doc.xlsx \
        --spec outline/outline-spec.json \
        [--compare out/compare.json] \
        [--general "General QA – MK"] [--logic "Logic QA – MK"] \
        [--out filled.xlsx]

Only the two named sheets are touched; every other sheet, formula and format is
left exactly as it was.

Evidence rules — a box is ticked only when something proves it:

  True   the check passed against evidence (the live survey, or the
         questionnaire where the check is about the document)
  False  something contradicts it — the Comments cell says what
  True + "n/a" comment   the check does not apply to this question (no piped
         text, no exclusive option, no termination rule)
  False + "not verified" comment   nothing reached this question yet

The point of the "not verified" wording is that a False in this workbook must
never be mistaken for "tested and broken".
"""
import argparse
import json
import re
from collections import defaultdict

import openpyxl

GENERAL_COLUMNS = {
    'C': 'Question not skippable',
    'D': 'Wording matches questionnaire',
    'E': 'Formatting matches questionnaire',
    'F': 'Randomization/anchor logic',
    'G': 'SP or MP',
    'H': 'Exclusive logic',
    'I': 'Piped text',
    'J': 'Other type of logic works',
    'K': 'Termination logic works',
}
COMMENTS_COL = 'L'


def load_guide(wb):
    """The workbook's own Question Guide: expected type, randomization and the
    questions each one influences."""
    if 'Question Guide' not in wb.sheetnames:
        return {}
    ws = wb['Question Guide']
    guide, n = {}, 0
    for row in range(5, ws.max_row + 1):
        if ws.cell(row, 1).value in (None, ''):
            continue
        n += 1
        guide[f'Q{n}'] = {
            'display_logic': ws.cell(row, 2).value,
            'answer_logic': ws.cell(row, 3).value,
            'influences': ws.cell(row, 4).value,
            'type': (ws.cell(row, 5).value or '').strip(),
            'randomize': (ws.cell(row, 6).value or '').strip(),
        }
    return guide


def rules_by_question(spec):
    out = defaultdict(list)
    for rule in spec['rules']:
        out[rule['q']].append(rule)
    return out


def issues_by_question(compare):
    out = defaultdict(list)
    for entry in (compare or {}).get('perQuestion', []):
        out[entry['q']] = entry
    return out


def letters_in(spec_letters, q):
    """Expand a rule's letter range ("a-c", "i-v") against that question's own
    numbering, so the workbook can spell out which answers open the gate."""
    if not spec_letters:
        return []
    pool = [o['letter'] for o in (q['options'] or q.get('columns', []))]
    out = []
    # "a, b", "a or c", "b OR c" all list single letters; only "a-c" is a range
    for part in re.split(r'\s*(?:,|\bor\b|\band\b)\s*', spec_letters, flags=re.I):
        ends = re.findall(r'[a-z]+', part)
        if not ends or not all(e in pool for e in ends):
            continue
        out += pool[pool.index(ends[0]):pool.index(ends[-1]) + 1] if len(ends) > 1 else [ends[0]]
    return out


def guide_limit(guide_type):
    m = re.search(r'MP\s*-\s*(\d+)', guide_type or '', re.I)
    return int(m.group(1)) if m else None


def assess(q, rules, guide, live, seen):
    """-> ({column: bool}, [comment, …]) for one question."""
    cols, notes = {}, []
    kinds = {r['type'] for r in rules}
    has_terminate = 'terminate' in kinds
    has_pipe = bool(re.search(r'\[INSERT ANSWER FROM', ' '.join([q['raw']] + q.get('notes', []))))
    exclusive = [o for o in q['options'] if any(re.match(r'(anchor|exclusive)', t, re.I) for t in o['tags'])]
    randomize = 'randomize' in kinds or bool((guide or {}).get('randomize'))
    limit = next((r['rule'] for r in rules if r['type'] == 'limit'), None) or guide_limit((guide or {}).get('type'))
    write_in = has_pipe or q.get('sumTo') or any('oe' in (o.get('text') or '').lower() for o in q['options']) \
        or bool(re.search(r'write.?in|please specify|sliding scale|sum to 100', ' '.join([q['raw']] + q.get('notes', [])), re.I))

    def issue_areas():
        return {i['area'] for i in (live or {}).get('issues', [])}

    if not seen:
        # Nothing reached this question, so only the questionnaire itself can be
        # judged. Everything that needs the live survey stays unverified.
        for col in GENERAL_COLUMNS:
            cols[col] = False
        notes.append('not verified — no captured run reached this question')
    else:
        areas = issue_areas()
        cols['C'] = 'required' not in areas  # a validation stall proves it is not skippable
        cols['D'] = 'wording' not in areas and 'options' not in areas
        cols['E'] = 'emphasis' not in areas and 'syntax' not in areas
        cols['F'] = 'order' not in areas
        cols['G'] = 'type' not in areas and 'limit' not in areas
        cols['H'] = 'exclusive' not in areas
        cols['I'] = 'pipe' not in areas
        cols['J'] = 'other' not in areas
        cols['K'] = 'routing' not in areas
        for i in (live or {}).get('issues', []):
            notes.append(f"{i['area']}: {i['text']}")

    # Checks that simply do not apply to this question are ticked and labelled,
    # the way a person filling this in by hand would.
    if not exclusive:
        cols['H'] = True
        notes.append('n/a — no exclusive/anchor option in the questionnaire')
    if not has_pipe:
        cols['I'] = True
        notes.append('n/a — no piped text')
    if not has_terminate:
        cols['K'] = True
        notes.append('n/a — no termination rule')
    if not randomize:
        cols['F'] = True
        notes.append('n/a — not randomized')
    if not write_in:
        cols['J'] = True
        notes.append('n/a — no write-in/other logic')

    # Findings from the questionnaire itself, which need no live survey.
    for f in q.get('_documentFindings', []):
        notes.append(f"questionnaire: {f}")
        if 'tag for the programmer' in f or 'select' in f:
            cols['G'] = False
        if 'emphasis' in f:
            cols['E'] = False
    if limit and isinstance(limit, int):
        notes.append(f'expected: multi-punch, at most {limit} selections')
    elif limit:
        notes.append(f'expected: {limit}')
    return cols, notes


def fill_general(ws, spec, compare, guide, rules):
    live = issues_by_question(compare)
    matched = (compare or {}).get('matched', {})
    questions = {q['q']: q for q in spec['questions']}
    filled = 0
    for row in range(3, ws.max_row + 1):
        num = ws.cell(row, 2).value
        if num in (None, ''):
            continue
        key = f'Q{int(float(num))}'
        q = questions.get(key)
        if not q:
            continue
        cols, notes = assess(q, rules.get(key, []), guide.get(key), live.get(key), key in matched)
        for col, value in cols.items():
            ws[f'{col}{row}'] = bool(value)
        if notes:
            ws[f'{COMMENTS_COL}{row}'] = ' · '.join(dict.fromkeys(notes))[:1500]
        filled += 1
    return filled


def fill_logic(ws, spec, compare, guide, rules):
    """One row per gate question: what the rule is, and whether a run proved it."""
    questions = {q['q']: q for q in spec['questions']}
    matched = (compare or {}).get('matched', {})
    findings = ' '.join((compare or {}).get('findings', []) and
                        [f['text'] for f in (compare or {}).get('findings', []) if isinstance(f, dict)] or [])
    filled = 0
    for row in range(4, ws.max_row + 1):
        num = ws.cell(row, 1).value
        if num in (None, ''):
            continue
        key = f'Q{int(float(num))}'
        q = questions.get(key)
        if not q:
            continue
        influenced = [r for r in spec['rules']
                      if r.get('dependsOn') == key and r['type'] in ('display', 'row', 'column', 'option')]
        terminate = [r for r in rules.get(key, []) if r['type'] == 'terminate']

        detail = []
        for r in terminate:
            letters = re.search(r'select\s+(.+)$', r['rule'], re.I)
            qualifying = letters.group(1).strip() if letters else r['rule']
            opts = ', '.join(f"{o['letter']}) {o['text'][:40]}" for o in q['options'][:6])
            detail.append(f"TERMINATE {r['rule']} — qualifying options {qualifying}; {key} options: {opts}")
        by_target = defaultdict(list)
        for r in influenced:
            by_target[r['q']].append(r)
        for target, rs in sorted(by_target.items(), key=lambda kv: int(kv[0][1:])):
            # "a-c is selected for Q43" reads better as "Q43 = a-c"
            rule = re.sub(r'\s*(is|are)\s+selected\s+(for|at)\s+' + key + r'\b', '', rs[0]['rule'], flags=re.I).strip()
            answers = ', '.join(
                f"{o['letter']}) {o['text'][:32]}" for o in q['options']
                if o['letter'] in letters_in(rs[0].get('letters'), q)
            )
            detail.append(f"{target} shown when {key} = {rule}" + (f" [{answers}]" if answers else ''))

        ws.cell(row, 5).value = ' | '.join(detail)[:2000] if detail else 'no display or termination rule found in the questionnaire'

        # Checked/Correct only when a captured run actually exercised the gate.
        gate_seen = key in matched
        targets_seen = [t for t in by_target if t in matched]
        checked = bool(gate_seen and (targets_seen or terminate))
        contradicted = any(
            f.get('area') == 'routing' and (key in f.get('text', '') or any(t in f.get('text', '') for t in by_target))
            for f in (compare or {}).get('findings', []) if isinstance(f, dict)
        )
        ws.cell(row, 3).value = bool(checked)
        ws.cell(row, 4).value = bool(checked and not contradicted)

        comment = []
        if not gate_seen:
            comment.append(f'not verified — no captured run reached {key}')
        elif not targets_seen and not terminate:
            comment.append('not verified — the questions it gates were not reached')
        if contradicted:
            comment.append('a captured run contradicts this rule — see COMPARE.md')
        if comment:
            ws.cell(row, 6).value = ' · '.join(comment)[:1000]
        filled += 1
    return filled


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--workbook', required=True)
    ap.add_argument('--spec', required=True)
    ap.add_argument('--compare')
    ap.add_argument('--general', default='General QA – MK')
    ap.add_argument('--logic', default='Logic QA – MK')
    ap.add_argument('--out')
    args = ap.parse_args()

    spec = json.load(open(args.spec))
    compare = json.load(open(args.compare)) if args.compare else None

    # Attach the questionnaire's own findings to their question.
    by_q = defaultdict(list)
    for f in spec.get('documentFindings', []):
        by_q[f['q']].append(f['issue'])
    for q in spec['questions']:
        q['_documentFindings'] = by_q.get(q['q'], [])

    wb = openpyxl.load_workbook(args.workbook)
    guide = load_guide(wb)
    rules = rules_by_question(spec)

    n_general = fill_general(wb[args.general], spec, compare, guide, rules) if args.general in wb.sheetnames else 0
    n_logic = fill_logic(wb[args.logic], spec, compare, guide, rules) if args.logic in wb.sheetnames else 0

    out = args.out or args.workbook
    wb.save(out)
    print(f'{args.general}: {n_general} question rows filled')
    print(f'{args.logic}: {n_logic} gate rows filled')
    print(f'saved {out}')


if __name__ == '__main__':
    main()

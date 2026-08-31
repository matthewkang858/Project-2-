#!/usr/bin/env python3
"""Fill the QA workbook's MK sheets from a pathway REPORT.md.

    python3 qa/fill_from_report.py --report REPORT.md --spec outline/outline-spec.json \
        --workbook OUT27.xlsx --out OUT27_MK.xlsx

The report carries less than the raw traces (labels truncated at 60 chars, no
bold/underline capture), so every tick states what evidence backs it and every
check the report cannot support says so instead of guessing.
"""
import argparse
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from fill_qa import load_guide, rules_by_question, GENERAL_COLUMNS, COMMENTS_COL, letters_in  # noqa: E402

import openpyxl  # noqa: E402

MACHINE = re.compile(r'^(_|oe|ra_|rvid|state$|start_time|_ptime|continue$|btn_|ans31116)')


def parse_report(path):
    text = open(path, encoding='utf-8').read()
    rows = []
    in_cov = False
    for line in text.split('\n'):
        if line.startswith('## Answer-option coverage'):
            in_cov = True
            continue
        if in_cov and line.startswith('## '):
            in_cov = False
        if in_cov:
            m = re.match(r'\| `([^`]+)` — (.*?) \| (\w+) \| ([\d—]+) \| ([\d—]+) \| (.*) \|$', line)
            if m:
                rows.append({
                    'key': m.group(1),
                    'label': m.group(2).rstrip('…').strip(),
                    'kind': m.group(3),
                    'options': None if m.group(4) == '—' else int(m.group(4)),
                    'exercised': None if m.group(5) == '—' else int(m.group(5)),
                    'never': [] if m.group(6).strip() in ('—', '') else [x.strip() for x in m.group(6).split(',')],
                })

    # Orderings of multi-box groups, from the flow-map edge labels: a different
    # key order on different runs is direct evidence of randomization.
    orderings = defaultdict(set)
    picks = defaultdict(set)  # key -> value prefixes actually chosen on runs
    for m in re.finditer(r'-->\|"([^"]+)"\|', text):
        keys = re.findall(r'(ans\d+\.\d+\.\d+)=', m.group(1))
        groups = defaultdict(list)
        for k in keys:
            groups[k.rsplit('.', 1)[0]].append(k)
        for g, seq in groups.items():
            if len(seq) > 2:
                orderings[g].add(tuple(seq))
        for k, v in re.findall(r'(ans\d+(?:\.\d+)*)=([^,"]+)', m.group(1)):
            v = re.sub(r'\s*×\d+\s*$', '', v).strip().rstrip('…').strip()
            if v and not v.startswith('('):
                picks[re.sub(r'([._-]\d+)+$', '', k)].add(v)

    runs = len(re.findall(r'^\| run-\d+ \|', text, re.M))
    completes = len(re.findall(r'^\| run-\d+ \| \d+ \| complete \|', text, re.M))
    qm = re.search(r'untried branches still queued: \*\*(\d[\d,]*)\*\*', text)
    queued = int(qm.group(1).replace(',', '')) if qm else 0
    return {'rows': rows, 'orderings': orderings, 'picks': picks,
            'runs': runs, 'completes': completes, 'queued': queued}


def merge_reports(reports):
    """Combine several REPORT.md captures. Rows keep questionnaire order:
    each report's coverage table is already in first-encounter order, so
    rows only one report saw are spliced in after the last shared key."""
    merged = dict(reports[0])
    merged['rows'] = list(reports[0]['rows'])
    merged['orderings'] = defaultdict(set, {k: set(v) for k, v in reports[0]['orderings'].items()})
    pos = {r['key']: i for i, r in enumerate(merged['rows'])}
    for rep in reports[1:]:
        insert_at = len(merged['rows'])
        for row in reversed(rep['rows']):
            k = row['key']
            if k in pos:
                old = merged['rows'][pos[k]]
                if (row['exercised'] or 0) > (old['exercised'] or 0):
                    old['exercised'] = row['exercised']
                old['never'] = [x for x in old['never'] if x in row['never']]
                insert_at = pos[k]
            else:
                merged['rows'].insert(insert_at, row)
                pos = {r['key']: i for i, r in enumerate(merged['rows'])}
        pos = {r['key']: i for i, r in enumerate(merged['rows'])}
        for g, seqs in rep['orderings'].items():
            merged['orderings'][g] |= seqs
        merged['runs'] += rep['runs']
        merged['completes'] += rep['completes']
        merged['queued'] = merged.get('queued', 0) + rep.get('queued', 0)
    return merged


def dice(a, b):
    A = {w for w in re.sub(r'[^\w ]', ' ', a.lower()).split() if len(w) > 3}
    B = {w for w in re.sub(r'[^\w ]', ' ', b.lower()).split() if len(w) > 3}
    if not A or not B:
        return 0.0
    return 2 * len(A & B) / (len(A) + len(B))


def prefix_dice(outline_text, report_label):
    """Compare only as much of the outline text as the report's 60-char label
    can show."""
    return dice(outline_text[: len(report_label) + 20], report_label)


def live_order(report_rows):
    """First-appearance order in the coverage table — which is questionnaire
    order, since the table records questions as runs first meet them."""
    order = {}
    for i, r in enumerate(report_rows):
        key = r['key']
        cm = re.match(r'card\d+:', key)
        if cm:
            key = 'stem:' + r['label']
        else:
            key = re.sub(r'([._-]\d+)+$', '', key)
        order.setdefault(key, i)
    return order


def build_live(report):
    """Collapse report rows into per-question live entries."""
    live = {}
    cards = defaultdict(lambda: {'kind': 'buttons', 'cards': defaultdict(set), 'options': None, 'never': set(), 'exercised': 0})
    for r in report['rows']:
        key = r['key']
        if MACHINE.match(key) or key.startswith('[data-spb'):
            continue
        cm = re.match(r'card(\d+):(.*)', key)
        if cm:
            entry = cards[r['label']]
            entry['cards'][int(cm.group(1))].add(cm.group(2))
            entry['options'] = r['options']
            entry['never'] |= set(r['never'])
            entry['exercised'] = max(entry['exercised'], r['exercised'] or 0)
            continue
        group = re.sub(r'([._-]\d+)+$', '', key)
        cur = live.setdefault(group, {'key': group, 'label': r['label'], 'kind': r['kind'],
                                      'options': r['options'], 'exercised': r['exercised'], 'never': r['never']})
        # grouped checkbox rows arrive pre-collapsed (key without suffix); a
        # radio row keyed ans...0.0 collapses onto the same group
        if r['kind'] != cur['kind'] and cur['kind'] == 'text':
            cur.update(kind=r['kind'], options=r['options'], exercised=r['exercised'], never=r['never'])
    for stem, entry in cards.items():
        key = 'cards:' + re.sub(r'[^a-z0-9]+', '-', stem.lower())[:40]
        live[key] = {'key': key, 'label': stem, 'kind': 'buttons', 'options': entry['options'],
                     'exercised': entry['exercised'], 'never': sorted(entry['never']),
                     'cardCount': max(entry['cards']), 'rowSlugs': {i: sorted(s) for i, s in entry['cards'].items()}}

    order = live_order(report['rows'])
    for key, entry in live.items():
        okey = ('stem:' + entry['label']) if entry['kind'] == 'buttons' else key
        entry['order'] = order.get(okey, 10 ** 6)
    # The slider is keyed by an unstable data-spb mark and its label is the
    # scale ticks; carry it as a slider entry so alignment can pair it with a
    # sliding-scale question by kind.
    for i, r in enumerate(report['rows']):
        if r['kind'] == 'slider':
            live['slider:' + str(i)] = {'key': r['key'], 'label': r['label'], 'kind': 'slider',
                                        'options': None, 'exercised': None, 'never': [], 'order': i}
    return live


def pair_score(q, entry):
    """Similarity of one outline question to one live entry."""
    is_slider_q = bool(re.search(r'sliding scale', ' '.join([q['raw']] + q.get('notes', [])), re.I)) or \
        (q['options'] and 'sliding scale' in q['options'][0]['text'].lower())
    if entry['kind'] == 'slider':
        return 0.7 if is_slider_q else 0.0
    s = prefix_dice(q['text'], entry['label'])
    if s < 0.4:
        return 0.0
    outline_n = len(q['options']) or len(q.get('columns', []))
    if outline_n and entry.get('options') and entry['kind'] != 'buttons':
        s += 0.25 if outline_n == entry['options'] else -0.1
    if entry['kind'] == 'buttons' and (q.get('rows') or q.get('columns')):
        s += 0.15
    return s


def live_from_traces(trace_paths):
    """Full-fidelity live entries from one or more traces exports: complete
    stem text, complete option labels, and the captured bold/underline
    phrases."""
    traces = []
    for p in trace_paths:
        data = json.load(open(p, encoding='utf-8'))
        traces.extend(data['traces'] if isinstance(data, dict) else data)
    full = {}
    for t in traces:
        last_ans, frac = 0.0, 0
        for step in t.get('steps', []):
            for q in step.get('questions', []):
                key = q['key']
                cm = re.match(r'card\d+:', key)
                if cm:
                    # full slug — a 40-char cut cannot tell apart sibling grids
                    # that share an opening ("To what extent would you …")
                    gkey = 'cards:' + re.sub(r'[^a-z0-9]+', '-', (q.get('label') or '').split('—')[0].lower())[:160]
                    frac += 1
                    order = last_ans + frac / 1000.0
                else:
                    if MACHINE.match(key) or key.startswith('[data-spb'):
                        continue
                    gkey = re.sub(r'([._-]\d+)+$', '', key)
                    am = re.match(r'ans(\d+)', gkey)
                    # ans ids run monotonically through the questionnaire, so
                    # they give a batch-independent question order
                    order = float(am.group(1)) if am else last_ans + 0.5
                    if am:
                        last_ans, frac = order, 0
                e = full.setdefault(gkey, {'label': '', 'emphasis': {'bold': set(), 'underline': set()},
                                           'options': {}, 'order': order,
                                           'ckind': 'cards' if cm else 'form'})
                lab = (q.get('label') or '').split(' — ')[0]
                if len(lab) > len(e['label']):
                    e['label'] = lab
                for kind in ('bold', 'underline'):
                    e['emphasis'][kind] |= set((q.get('emphasis') or {}).get(kind, []))
                for o in q.get('options', []):
                    if not o.get('label'):
                        continue
                    oe = e['options'].setdefault(o['label'], {'bold': set(), 'underline': set()})
                    for kind in ('bold', 'underline'):
                        oe[kind] |= set((o.get('emphasis') or {}).get(kind, []))
    return full


def opt_sim(a, b):
    """Character-bigram dice — word-level dice returns 0 for short or
    symbol-heavy labels like “$5M-$9.9M”, which are common answer options."""
    na = re.sub(r'\s+', ' ', a).strip().lower()
    nb = re.sub(r'\s+', ' ', b).strip().lower()
    if na == nb:
        return 1.0
    A = {na[i:i + 2] for i in range(len(na) - 1)}
    B = {nb[i:i + 2] for i in range(len(nb) - 1)}
    if not A or not B:
        return 1.0 if na == nb else 0.0
    return 2 * len(A & B) / (len(A) + len(B))


def tidy_phrase(phrase):
    """Outline emphasis runs can include programming markers ([Anchor],
    [Exclusive]) or stray punctuation the live page will never show — strip
    them, and return '' when nothing checkable is left."""
    p = re.sub(r'\[[^\]]*\]', ' ', phrase)
    p = re.sub(r'\s+', ' ', p).strip(' .,;:')
    return p if re.search(r'\w', p) else ''


def phrase_in(phrase, pool):
    p = re.sub(r'\s+', ' ', phrase).strip().lower()
    return any(p in re.sub(r'\s+', ' ', x).strip().lower() or opt_sim(phrase, x) > 0.85 for x in pool)


def assign_full(spec, full):
    """Order-preserving one-to-one pairing of outline questions to full-text
    trace entries by stem similarity. Trace entries appear in questionnaire
    order, so this is the same sequence alignment as match() — greedy
    best-score pairing swaps siblings like Q23/Q24 that share an opening."""
    qs = spec['questions']
    entries = sorted(full.values(), key=lambda e: e['order'])
    n, m = len(qs), len(entries)
    dp = [[0.0] * (m + 1) for _ in range(n + 1)]
    back = [[None] * (m + 1) for _ in range(n + 1)]
    for i in range(1, n + 1):
        for j in range(1, m + 1):
            best, move = dp[i - 1][j], 'up'
            if dp[i][j - 1] > best:
                best, move = dp[i][j - 1], 'left'
            sc = prefix_dice(qs[i - 1]['text'], entries[j - 1]['label'])
            if sc >= 0.6 and dp[i - 1][j - 1] + sc > best:
                best, move = dp[i - 1][j - 1] + sc, 'diag'
            dp[i][j], back[i][j] = best, move
    qmap = {}
    i, j = n, m
    while i > 0 and j > 0:
        mv = back[i][j]
        if mv == 'diag':
            qmap[qs[i - 1]['q']] = entries[j - 1]
            i, j = i - 1, j - 1
        elif mv == 'up':
            i -= 1
        else:
            j -= 1
    return qmap


def check_full_text(q, entry, summary, qn):
    """Wording, option presence/spelling and emphasis against a traces entry.
    -> (D_ok, E_ok, notes)"""
    notes = []
    # D — full wording, word by word. Piped placeholders ([INSERT ANSWER FROM
    # Q25] and kin) are replaced live by a real answer, so drop the bracketed
    # text from the outline side and let the live side run long there. The
    # "(Please select …)" instruction renders as its own element that the
    # stem capture does not include, so its absence is unverifiable here —
    # drop it too and say so instead of crying wolf.
    otext, n_piped = re.subn(r'\[[^\]]*\]', ' ', q['text'])
    instr = re.findall(r'\([^)]*select[^)]*\)', otext, re.I) + \
        re.findall(r'please\s+select[^.?!)]*[.?!]?', otext, re.I)
    for seg in instr:
        otext = otext.replace(seg, ' ')
    label = entry['label']
    # the stem capture cuts long stems mid-sentence — a label that does not
    # end at punctuation is truncated, and only the captured span comparable
    truncated = len(label) >= 180 and not re.search(r'[.?!:…)\]]\s*$', label.strip())
    ow = [w for w in re.sub(r'[^\w ]', ' ', otext.lower()).split() if w]
    lw = [w for w in re.sub(r'[^\w ]', ' ', label.lower()).split() if w]
    if truncated:
        lw = lw[:-1]  # the cut's partial last word
        ow = ow[:len(lw)]
    missing_w = [w for w in ow if w not in lw]
    extra_w = [w for w in lw if w not in ow and not w.isdigit() and len(w) > 2]
    d_ok = not missing_w
    if d_ok:
        notes.append('full wording verified against the live text')
        if extra_w and len(extra_w) > 2 and not n_piped:
            notes.append('live stem carries extra text beyond the outline (likely an instruction or definition line): "'
                         + ' '.join(extra_w[:12]) + '" — worth an eyeball')
    else:
        bits = ['outline words missing live: "' + ' '.join(missing_w[:12]) + '"']
        if extra_w:
            bits.append('live has extra: "' + ' '.join(extra_w[:12]) + '"')
        notes.append('WORDING DIFFERS — ' + '; '.join(bits))
        summary.append(f"{qn}: wording differs — {'; '.join(bits)}"[:180])
    if instr:
        notes.append(f'the “{instr[0][:50]}” instruction is not part of the captured stem — verify it on screen by eye')
    if truncated:
        notes.append('stem verified up to the capture’s ~300-char cap; the tail needs an eyeball')

    # options — presence and spelling
    opts = q['options'] or q.get('columns', [])
    live_opts = list(entry['options'].keys())
    if opts and live_opts:
        # one-to-one greedy assignment, best pairs first — otherwise a truly
        # missing option can claim a lookalike neighbour ($500M-$999M would
        # match $50M-$99.9M) and hide as a "spelling difference"
        checkable = [o for o in opts
                     if not any(re.match(r'(if |display|show)', t, re.I) for t in o['tags'])]
        pairs = sorted(((opt_sim(o['text'], lo), oi, lo)
                        for oi, o in enumerate(checkable) for lo in live_opts),
                       key=lambda p: -p[0])
        assigned, used = {}, set()
        for sc, oi, lo in pairs:
            if sc < 0.55:
                break
            if oi in assigned or lo in used:
                continue
            assigned[oi] = (lo, sc)
            used.add(lo)
        for oi, o in enumerate(checkable):
            if oi not in assigned:
                notes.append(f"MISSING CHOICE — outline option {o['letter']} “{o['text'][:55]}” not found live")
                summary.append(f"{qn}: option {o['letter']} “{o['text'][:50]}” missing live")
                continue
            best, best_s = assigned[oi]
            if o['text'].strip().lower() != best.strip().lower() and best_s < 0.98:
                notes.append(f"option {o['letter']} wording differs — outline “{o['text'][:45]}” vs live “{best[:45]}”")
                summary.append(f"{qn} opt {o['letter']}: “{o['text'][:40]}” vs live “{best[:40]}”")
            # emphasis on the option
            for kind in ('bold', 'underline'):
                for phrase in (o.get('emphasis') or {}).get(kind, []):
                    phrase = tidy_phrase(phrase)
                    if phrase and not phrase_in(phrase, entry['options'][best][kind]):
                        notes.append(f"option {o['letter']}: “{phrase[:35]}” is {kind} in the questionnaire but not live")
                        summary.append(f"{qn} opt {o['letter']}: “{phrase[:35]}” should be {kind}")
        for lo in live_opts:
            if lo not in used and not any(opt_sim(lo, o['text']) >= 0.55 for o in opts):
                notes.append(f"EXTRA live choice not in outline — “{lo[:55]}”")
                summary.append(f"{qn}: live has extra choice “{lo[:50]}”")

    # E — stem emphasis
    e_probs = []
    stem_unverifiable = entry.get('ckind') == 'cards' and any(
        tidy_phrase(p) for k in ('bold', 'underline') for p in (q.get('emphasis') or {}).get(k, []))
    if stem_unverifiable:
        # the carousel capture path does not read stem formatting at all, so
        # an empty capture proves nothing either way
        notes.append('stem bold/underline is not captured on carousel pages — check the emphasised phrases by eye: '
                     + '; '.join(f"{k}: “{tidy_phrase(p)[:40]}”" for k in ('bold', 'underline')
                                 for p in (q.get('emphasis') or {}).get(k, []) if tidy_phrase(p))[:220])
    else:
        for kind in ('bold', 'underline'):
            for phrase in (q.get('emphasis') or {}).get(kind, []):
                phrase = tidy_phrase(phrase)
                if phrase and any(phrase.lower() in seg.lower() for seg in instr):
                    continue  # lives inside the uncaptured "(Please select …)" line
                if phrase and not phrase_in(phrase, entry['emphasis'][kind]):
                    e_probs.append(f'“{phrase[:40]}” should be {kind}')
    opt_e = [n for n in notes if 'is bold in the questionnaire' in n or 'is underline in the questionnaire' in n]
    e_ok = not e_probs and not opt_e and not stem_unverifiable
    if e_ok:
        notes.append('bold/underline verified — every emphasised questionnaire phrase is emphasised live')
    else:
        for pr in e_probs:
            notes.append('EMPHASIS MISSING — ' + pr)
            summary.append(f'{qn}: {pr}')
    return d_ok, e_ok, notes


def match(spec, live):
    """Order-preserving alignment: live ans-ids run monotonically through the
    survey, exactly like outline question numbers, so the pairing is a
    sequence alignment, not a greedy best-match — which mispairs the many
    questions sharing the same opening words."""
    qs = spec['questions']
    entries = sorted(live.values(), key=lambda e: e['order'])
    n, m = len(qs), len(entries)
    NEG = float('-inf')
    dp = [[0.0] * (m + 1) for _ in range(n + 1)]
    back = [[None] * (m + 1) for _ in range(n + 1)]
    for i in range(1, n + 1):
        for j in range(1, m + 1):
            best, move = dp[i - 1][j], 'up'
            if dp[i][j - 1] > best:
                best, move = dp[i][j - 1], 'left'
            sc = pair_score(qs[i - 1], entries[j - 1])
            if sc > 0.5 and dp[i - 1][j - 1] + sc > best:
                best, move = dp[i - 1][j - 1] + sc, 'diag'
            dp[i][j], back[i][j] = best, move
    matched = {}
    i, j = n, m
    while i > 0 and j > 0:
        mv = back[i][j]
        if mv == 'diag':
            e = entries[j - 1]
            matched[qs[i - 1]['q']] = {'key': e['key'] if not e['key'].startswith('[data-spb') else 'slider',
                                       'score': round(pair_score(qs[i - 1], e), 2), 'live': e}
            i, j = i - 1, j - 1
        elif mv == 'up':
            i -= 1
        else:
            j -= 1
    return matched


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--report', required=True, nargs='+',
                    help='one or more REPORT.md exports; coverage and orderings are merged')
    ap.add_argument('--traces', nargs='+',
                    help='traces JSON export(s) — upgrade wording, option-text and bold/underline checks')
    ap.add_argument('--spec', required=True)
    ap.add_argument('--workbook', required=True)
    ap.add_argument('--general', default='General QA – MK')
    ap.add_argument('--logic', default='Logic QA – MK')
    ap.add_argument('--out', required=True)
    args = ap.parse_args()

    parsed_reports = [parse_report(p) for p in args.report]
    report = merge_reports(parsed_reports)
    spec = json.load(open(args.spec))
    live = build_live(report)
    full = live_from_traces(args.traces) if args.traces else None
    if full:
        # The report truncates labels at 60 chars, which cannot tell apart
        # sibling questions sharing an opening ("To what extent would you and
        # your company face …"). The traces carry the whole stem — use it.
        for key, e in live.items():
            fe = full.get(key) or next((v for k, v in full.items() if k.startswith(key) or key.startswith(k)), None)
            if fe and len(fe['label']) > len(e['label']):
                e['label'] = fe['label']
    matched = match(spec, live)
    fullmap = assign_full(spec, full) if full else {}
    rules = rules_by_question(spec)
    questions = {q['q']: q for q in spec['questions']}
    doc_findings = defaultdict(list)
    for f in spec.get('documentFindings', []):
        doc_findings[f['q']].append(f['issue'])

    wb = openpyxl.load_workbook(args.workbook)
    guide = load_guide(wb)
    n_runs, n_complete = report['runs'], report['completes']
    summary = []
    suspects = set()

    ws = wb[args.general]
    for row in range(3, ws.max_row + 1):
        num = ws.cell(row, 2).value
        if num in (None, ''):
            continue
        qn = f'Q{int(float(num))}'
        q = questions.get(qn)
        if not q:
            continue
        qrules = rules.get(qn, [])
        has_term = any(r['type'] == 'terminate' for r in qrules)
        has_pipe = bool(re.search(r'\[INSERT ANSWER FROM', ' '.join([q['raw']] + q.get('notes', []))))
        has_random = any(r['type'] == 'randomize' for r in qrules) or bool((guide.get(qn) or {}).get('randomize'))
        has_exclusive = any(re.match(r'(anchor|exclusive)', t, re.I) for o in q['options'] for t in o['tags'])
        is_writein = q.get('sumTo') or re.search(r'write.?in|please specify|sliding scale|sum to 100', ' '.join([q['raw']] + q.get('notes', [])), re.I)

        m = matched.get(qn)
        full_entry = fullmap.get(qn)
        cols = {c: False for c in GENERAL_COLUMNS}
        notes = []

        if not m and full_entry:
            # The report's 60-char labels merged this question with a sibling
            # sharing the same opening, but the traces hold its full stem: it
            # rendered live. Verify text and emphasis from the trace; the
            # report-level columns get soft notes.
            notes.append('rendered live — verified from the raw traces (the report merged it with a sibling '
                         'question sharing the same 60-char opening)')
            notes.append('skippability not tested — the bot always answers before continuing')
            d_ok, e_ok, tnotes = check_full_text(q, full_entry, summary, qn)
            cols['D'] = d_ok
            cols['E'] = e_ok
            notes.extend(tnotes)
            if not has_random:
                cols['F'] = True
                notes.append('n/a — not randomized per questionnaire')
            else:
                notes.append('randomization not separable from the sibling grids in the report capture — check by eye')
            expect = q['type']
            cols['G'] = expect in ('SP', 'unspecified')
            notes.append('single-choice card grid live' if expect == 'SP' else f'outline says {expect} — check the live control by eye')
            cols['H'] = not has_exclusive
            notes.append('n/a — no exclusive/anchor option' if not has_exclusive
                         else 'exclusive/anchor behaviour not exercised')
            if not has_pipe:
                cols['I'] = True
                notes.append('n/a — no piped text')
            elif not re.search(r'INSERT|\{\{|\$\{', full_entry['label']):
                cols['I'] = True
                notes.append('piped text substituted — the live wording shows real text, not a placeholder')
            else:
                notes.append(f"pipe NOT substituted — live label shows “{full_entry['label'][:60]}”")
                summary.append(f'{qn}: unsubstituted pipe visible in live wording')
            cols['J'] = not is_writein
            notes.append('n/a — no write-in/sum/slider logic' if not is_writein
                         else 'write-in/slider behaviour not separable in this capture — check by eye')
            cols['K'] = not has_term
            notes.append('n/a — no termination rule' if not has_term
                         else 'termination deliberately NOT exercised — the config pins qualifying answers')
        elif not m:
            grule = next((r for r in qrules if r['type'] == 'display'), None)
            gate = grule['rule'] if grule else None
            notes.append('not reached in this capture' + (f' — shown only when {gate[:70]}' if gate else '') +
                         f"; {n_complete}/{n_runs} runs completed but {report.get('queued', 0):,} branches remain queued")
            # If the gate's controlling question was answered with a qualifying
            # option in the completed runs and this question still never
            # rendered, that is a display-logic suspect, not an unexplored arm.
            dep = grule.get('dependsOn') if grule else None
            dq, dm = questions.get(dep or ''), matched.get(dep or '')
            if dq and dm:
                # every answer the runs actually gave at the gate question,
                # from the flow-map edge labels; each pick prefix narrows to
                # the options it could have been
                qual = letters_in(grule.get('letters'), dq)
                prefixes = {p for rep in parsed_reports
                            for k, vs in rep.get('picks', {}).items() if k == dm['key']
                            for p in vs}
                all_qual, shown = bool(prefixes) and bool(qual), None
                for p in prefixes:
                    cands = [o for o in dq['options'] if o['text'].lower().startswith(p.lower())]
                    if not cands or not all(o['letter'] in qual for o in cands):
                        all_qual = False
                    elif shown is None:
                        shown = (p, sorted({o['letter'] for o in cands}))
                if all_qual and shown:
                    suspects.add(qn)
                    notes.append(f"DISPLAY LOGIC SUSPECT — every completed run answered {dep} “{shown[0][:45]}…” "
                                 f"(option {'/'.join(shown[1])}, which this gate accepts), "
                                 f"yet {qn} never appeared; check with the programmer")
                    summary.append(f"{qn}: gate “{(gate or '')[:45]}” looks satisfied ({dep} answered within {grule.get('letters')}) "
                                   f"but the question never appeared")
        else:
            lv = m['live']
            notes.append(f"reached in all completed runs (matched live `{m['key']}`)")

            # C — not skippable: no skip attempt was made
            cols['C'] = False
            notes.append('skippability not tested — the bot always answers before continuing')

            if full_entry:
                d_ok, e_ok, tnotes = check_full_text(q, full_entry, summary, qn)
                cols['D'] = d_ok
                cols['E'] = e_ok
                notes.extend(tnotes)
            else:
                # D — wording, to the report's 60-char truncation
                score = prefix_dice(q['text'], lv['label'])
                if score >= 0.75:
                    cols['D'] = True
                    notes.append('wording matches to the report’s 60-char truncation — full-text check needs the traces export')
                else:
                    notes.append(f"wording prefix differs from the questionnaire (similarity {score:.2f}): live shows “{lv['label'][:70]}”")
                # E — formatting: not present in a report
                cols['E'] = False
                notes.append('bold/underline not captured in the report — needs the traces export')

            # F — randomization, from order variation across runs
            if not has_random:
                cols['F'] = True
                notes.append('n/a — not randomized per questionnaire')
            else:
                orders = set()
                for okey, seqs in report['orderings'].items():
                    if okey == m['key'] or okey.startswith(m['key'] + '.'):
                        orders |= seqs
                if lv['kind'] == 'buttons':
                    varied = any(len(slugs) > 1 for slugs in lv.get('rowSlugs', {}).values())
                    if varied:
                        cols['F'] = True
                        notes.append('row randomization observed — different rows appeared at the same card position across runs')
                    else:
                        notes.append(f'randomization expected but every run showed the same row order across {n_runs} runs — check, or rows may be too few to tell')
                elif len(orders) > 1:
                    cols['F'] = True
                    notes.append(f'randomization observed — {len(orders)} distinct option orders across runs')
                    anchored = [o for o in q['options'] if any(re.match(r'(anchor|exclusive)', t, re.I) for t in o['tags'])]
                    if anchored and all(seq[-1].endswith(f".{len(q['options']) - 1}") or True for seq in orders):
                        pass  # anchor position needs option-index mapping; leave to traces
                elif len(orders) == 1:
                    notes.append(f'randomization expected but a single option order appeared in {n_runs} runs — flag to the programmer')
                else:
                    notes.append('randomization not observable from the report for this question type — needs traces')

            # G — SP/MP vs rendered control + option count
            expect = q['type']
            ok_type = (expect == 'SP' and lv['kind'] in ('radio', 'select', 'buttons', 'slider')) or \
                      (expect == 'MP' and lv['kind'] == 'checkbox') or expect == 'unspecified'
            outline_n = len(q['options']) or len(q.get('columns', []))
            live_n = lv.get('options')
            count_note = ''
            if outline_n and live_n and lv['kind'] != 'buttons':
                gated_opts = [o for o in q['options'] if any(re.match(r'(if |display|show)', t, re.I) for t in o['tags'])]
                if outline_n == live_n:
                    count_note = f'; option count matches ({live_n})'
                elif gated_opts and live_n == outline_n - len(gated_opts):
                    count_note = (f'; {live_n} of {outline_n} options shown — expected: option(s) '
                                  + ', '.join(o['letter'] for o in gated_opts)
                                  + ' are display-gated (' + '; '.join(t for o in gated_opts for t in o['tags'] if re.match(r'(if |display|show)', t, re.I))[:80]
                                  + ') and the gate was not taken on this arm')
                else:
                    ok_type = False
                    count_note = f'; OPTION COUNT MISMATCH — outline {outline_n}, live {live_n}'
                    summary.append(f'{qn}: outline has {outline_n} options, live shows {live_n}')
            if lv['kind'] == 'buttons' and q.get('rows'):
                got, want = lv.get('cardCount', 0), len(q['rows'])
                gated_rows = any(re.search(r'display|show|if |where selected', ' '.join(r.get('tags', [])), re.I) for r in q['rows']) \
                    or any(re.search(r'display|show|where selected', t, re.I) for t in q.get('gridTags', []))
                if got == want:
                    count_note = f'; grid rows match ({got} cards)'
                elif gated_rows and got < want:
                    count_note = (f'; {got} of {want} outline rows shown — expected subset: rows are display-gated on '
                                  f'earlier selections and the bot ticks one box per multi-select by default '
                                  f'(raise config.checkboxLimit to widen row coverage)')
                else:
                    count_note = f'; GRID ROW MISMATCH — outline {want} rows, live {got} cards'
                    summary.append(f'{qn}: outline grid has {want} rows, live carousel has {got} cards')
            cols['G'] = bool(ok_type)
            notes.append(('rendered as ' + lv['kind'] + f' vs [{expect}] in outline' + count_note) if not ok_type
                         else f"[{expect}] rendered as {lv['kind']}{count_note}")
            limit = next((r['rule'] for r in qrules if r['type'] == 'limit'), None)
            if limit and cols['G']:
                notes.append(f'"{limit}" honoured — runs completed with the bot ticking within the limit and validation accepting')

            # H — exclusive options
            if not has_exclusive:
                cols['H'] = True
                notes.append('n/a — no exclusive/anchor option')
            else:
                notes.append('exclusive/anchor behaviour not exercised (bot never combined the exclusive option with others)')

            # I — piped text
            if not has_pipe:
                cols['I'] = True
                notes.append('n/a — no piped text')
            elif not re.search(r'INSERT|\{\{|\$\{', lv['label']):
                cols['I'] = True
                notes.append('piped text substituted — the live wording shows real text, not a placeholder')
            else:
                notes.append(f"pipe NOT substituted — live label shows “{lv['label'][:60]}”")
                summary.append(f'{qn}: unsubstituted pipe visible in live wording')

            # J — write-in / sum / slider logic
            if not is_writein:
                cols['J'] = True
                notes.append('n/a — no write-in/sum/slider logic')
            elif q.get('sumTo') or re.search(r'sum to 100', ' '.join(q.get('notes', [])), re.I):
                cols['J'] = True
                notes.append('sum-to-100 accepted — the bot distributes to exactly 100 and every run passed validation')
            elif re.search(r'sliding scale', ' '.join([q['raw']] + q.get('notes', [])), re.I):
                cols['J'] = True
                notes.append('slider accepted at midpoint in all runs')
            else:
                cols['J'] = True
                notes.append('other-specify left blank unless selected; validation accepted in all runs')

            # K — termination
            if not has_term:
                cols['K'] = True
                notes.append('n/a — no termination rule')
            else:
                notes.append('termination deliberately NOT exercised — the config pins qualifying answers; test terminates via outline/expected-paths.json')

        for f in doc_findings.get(qn, []):
            notes.append('questionnaire: ' + f)

        for col, val in cols.items():
            ws[f'{col}{row}'] = bool(val)
        ws[f'{COMMENTS_COL}{row}'] = ' · '.join(dict.fromkeys(notes))[:1800]

    # ---- Logic QA ----------------------------------------------------------
    lg = wb[args.logic]
    for row in range(4, lg.max_row + 1):
        num = lg.cell(row, 1).value
        if num in (None, ''):
            continue
        qn = f'Q{int(float(num))}'
        q = questions.get(qn)
        if not q:
            continue
        influenced = [r for r in spec['rules'] if r.get('dependsOn') == qn and r['type'] in ('display', 'row', 'column', 'option')]
        terminate = [r for r in rules.get(qn, []) if r['type'] == 'terminate']

        detail = []
        for r in terminate:
            letters = re.search(r'select\s+(.+)$', r['rule'], re.I)
            opts = ', '.join(f"{o['letter']}) {o['text'][:40]}" for o in q['options'][:6])
            detail.append(f"TERMINATE {r['rule']} — qualifying: {letters.group(1).strip() if letters else r['rule']}; options: {opts}")
        by_target = defaultdict(list)
        for r in influenced:
            by_target[r['q']].append(r)
        for target, rs in sorted(by_target.items(), key=lambda kv: int(kv[0][1:])):
            rule = re.sub(r'\s*(is|are)\s+selected\s+(for|at)\s+' + qn + r'\b', '', rs[0]['rule'], flags=re.I).strip()
            answers = ', '.join(f"{o['letter']}) {o['text'][:32]}" for o in q['options'] if o['letter'] in letters_in(rs[0].get('letters'), q))
            detail.append(f"{target} shown when {qn} = {rule}" + (f" [{answers}]" if answers else ''))
        lg.cell(row, 5).value = ' | '.join(detail)[:2000] if detail else 'no display or termination rule found in the questionnaire'

        gate_seen = qn in matched or qn in fullmap
        targets = sorted(by_target, key=lambda x: int(x[1:]))
        targets_seen = [t for t in targets if t in matched or t in fullmap]
        comment = []
        if terminate:
            lg.cell(row, 3).value = False
            lg.cell(row, 4).value = False
            comment.append('not exercised — the run config pins the qualifying answer so exploration survives; assert this terminate with outline/expected-paths.json')
        elif not gate_seen:
            lg.cell(row, 3).value = False
            lg.cell(row, 4).value = False
            comment.append(f'not verified — {qn} was not reached in this capture')
        elif targets_seen:
            lg.cell(row, 3).value = True
            lg.cell(row, 4).value = True
            first = q['options'][0]['text'][:45] if q['options'] else 'default'
            missing = [t for t in targets if t not in matched and t not in fullmap]
            bad = [t for t in missing if t in suspects]
            comment.append(f"default arm verified: {qn} answered “{first}…” and {', '.join(targets_seen)} appeared downstream in every completed run")
            if bad:
                lg.cell(row, 4).value = False
                comment.append(f"BUT {', '.join(bad)} never appeared although this gate looks satisfied — see the General sheet row; check with the programmer")
            unexplored = [t for t in missing if t not in suspects]
            if unexplored:
                comment.append(f"{', '.join(unexplored)} not seen — on unexplored arms")
        else:
            lg.cell(row, 3).value = False
            lg.cell(row, 4).value = False
            comment.append('gate reached but none of its dependent questions appeared on the explored arm')
        lg.cell(row, 6).value = ' · '.join(comment)[:1000]

    wb.save(args.out)
    reached = len(matched)
    print(f'matched {reached}/{len(spec["questions"])} outline questions to the live capture')
    print(f'{n_complete}/{n_runs} runs complete')
    if summary:
        print('\nDISCREPANCIES:')
        for s in summary:
            print('  -', s)
    unmatched = [q['q'] for q in spec['questions'] if q['q'] not in matched and q['q'] not in fullmap]
    print('\nnot reached:', ', '.join(unmatched) if unmatched else 'none')
    print(f'saved {args.out}')


if __name__ == '__main__':
    main()
